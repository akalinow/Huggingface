import pandas as pd

##############################################
##############################################
def getKGmap(fileName, keyColumn='Konto KG', valueColumn='Opis SAP'):
    import pandas as pd
    # Load the mapping file (assuming it's an Excel file)
    # set NaN to "unset" 
    mapping_df = pd.read_excel(fileName)
    mapping_df[valueColumn] = mapping_df[valueColumn].fillna("unset")   
    
    # Create a dictionary to map keys to descriptions
    value_map = dict(zip(mapping_df[keyColumn], mapping_df[valueColumn]))
    
    return value_map
##############################################
##############################################
def getKGmaps(fileName, keyColumn='Konto KG'):
    kgToSAP = getKGmap(fileName, keyColumn=keyColumn, valueColumn="Opis SAP") 
    kgToNCNCat1 = getKGmap(fileName, keyColumn=keyColumn, valueColumn="NCN Kategoria I")
    kgToNCNCat2 = getKGmap(fileName, keyColumn=keyColumn, valueColumn="NCN Kategoria II")
    kgToNCNCat3 = getKGmap(fileName, keyColumn=keyColumn, valueColumn="NCN Kategoria III")

    return kgToSAP, kgToNCNCat1, kgToNCNCat2, kgToNCNCat3

##############################################
##############################################
def loadSAPData(fileName, mapFileName):
    import pandas as pd
    df = pd.read_excel(fileName)

    keyColumn = 'Konto KG'
    kgToSAP, kgToNCNCat1, kgToNCNCat2, kgToNCNCat3 = getKGmaps(mapFileName) 
    df['NCN Kategoria I'] = df[keyColumn].map(kgToNCNCat1)
    df['NCN Kategoria II'] = df[keyColumn].map(kgToNCNCat2)
    df['NCN Kategoria III'] = df[keyColumn].map(kgToNCNCat3)
    
    return df
##############################################
##############################################
def getOtherCostsTab(df, yearColumn='Rok obrotowy', categoryColumn='NCN Kategoria I', 
                     valueColumn='Kwota w WKr'):
  
    result = df.groupby([categoryColumn, yearColumn])[valueColumn].sum().unstack()
    result = result.fillna(0)

    #sort categories in custom order
    custom_order = ['koszty nab. aparatury specj.zalicz do ŚT',
                    'materiały', 
                    'pozostałe usł',
                    'Podróże służbowe',
                    'koszty konferencji i seminariów']
    result = result.reindex(custom_order, fill_value=0)

    # rename index categories to more readable names
    category_mapping = {
        'koszty nab. aparatury specj.zalicz do ŚT': 'Aparatura',
        'materiały': 'Materiały i drobny sprzęt',
        'pozostałe usł': 'Usługi obce',
        'koszty konferencji i seminariów': 'Konferencje i seminaria'
    }
    result = result.rename(index=category_mapping)

    result.index.name = ' '
    return result
###############################################
###############################################
def writeNCNReport(df, inputFileName):

    outputFileName = inputFileName.replace('.xlsx', '_report.xlsx')
    import pandas as pd
    with pd.ExcelWriter(outputFileName, engine='openpyxl') as writer:

        # Write your grouped summary to the second sheet
        otherCostsTab = getOtherCostsTab(df)
        otherCostsTab.to_excel(writer, sheet_name='Summary')

        # Write the main data to the first sheet
        df.to_excel(writer, sheet_name='SAP', index=False)

        worksheet = writer.sheets['Summary']
        from openpyxl.utils import get_column_letter

        
        for i, col in enumerate(otherCostsTab.columns, 1):
            # Find the maximum length in the column (header vs data)
            column_len = max(otherCostsTab[col].astype(str).map(len).max(), len(str(col))) + 2
            worksheet.column_dimensions[get_column_letter(i)].width = column_len
        
        index_len = max(otherCostsTab.index.astype(str).map(len).max(), len(str(otherCostsTab.index.name))) + 2
        worksheet.column_dimensions[get_column_letter(1)].width = index_len


    print("File saved successfully!")
    ################################################
    ################################################